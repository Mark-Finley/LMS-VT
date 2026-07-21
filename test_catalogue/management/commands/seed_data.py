from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from test_catalogue.models import TestCategory, Test, TestParameter
from django.db import transaction

User = get_user_model()

class Command(BaseCommand):
    help = 'Seeds initial test data, categories, and user accounts for all laboratory roles'

    @transaction.atomic
    def handle(self, *args, **options):
        self.stdout.write('Seeding user accounts...')
        
        # Define roles to create
        roles = [
            (User.ROLE_ADMINISTRATOR, 'admin', 'admin@lms.com'),
            (User.ROLE_LAB_MANAGER, 'manager', 'manager@lms.com'),
            (User.ROLE_PATHOLOGIST, 'pathologist', 'pathologist@lms.com'),
            (User.ROLE_LAB_SCIENTIST, 'scientist', 'scientist@lms.com'),
            (User.ROLE_LAB_TECHNICIAN, 'technician', 'technician@lms.com'),
            (User.ROLE_RECEPTIONIST, 'receptionist', 'receptionist@lms.com'),
            (User.ROLE_CASHIER, 'cashier', 'cashier@lms.com'),
            (User.ROLE_STORE_OFFICER, 'store', 'store@lms.com'),
            (User.ROLE_QUALITY_OFFICER, 'quality', 'quality@lms.com'),
            (User.ROLE_DOCTOR, 'doctor', 'doctor@lms.com'),
        ]

        for role, username, email in roles:
            if not User.objects.filter(username=username).exists():
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password='Password123!',
                    role=role,
                    first_name=username.capitalize(),
                    last_name='Staff'
                )
                self.stdout.write(self.style.SUCCESS(f"Created {role} user: {username}"))
            else:
                self.stdout.write(f"User {username} already exists.")

        self.stdout.write('Dynamically seeding test catalogue from Lab_Results_Template.xlsx...')
        import openpyxl
        import re
        import os
        from django.conf import settings
        excel_path = os.path.join(settings.BASE_DIR, "Lab_Results_Template.xlsx")
        if not os.path.exists(excel_path):
            excel_path = r"c:\Users\HP\Desktop\LMS\Lab_Results_Template.xlsx"

        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
            ws = wb['Lab Results']
            
            # Keep track of active test object we are building parameters for
            current_test = None
            current_category = None

            # Helper to map Excel test header to a system category
            def get_category_for_test(test_name):
                test_name_lower = test_name.lower()
                if any(x in test_name_lower for x in ['liver', 'kidney', 'glucose', 'lipid', 'thyroid', 'csf biochemical', 'iron', 'vitamin', 'biochemical', 'rft', 'kft']):
                    return 'Biochemistry'
                if any(x in test_name_lower for x in ['blood count', 'fbc', 'cbc', 'haemoglobin', 'g6pd', 'electrophoresis', 'coagulation', 'grouping']):
                    return 'Haematology'
                if any(x in test_name_lower for x in ['prolactin', 'oestrogen', 'estradiol', 'progesterone', 'testosterone', 'amh', 'fsh', 'hcg', 'hormone']):
                    return 'Hormones'
                if any(x in test_name_lower for x in ['afp', 'alpha fetoprotein']):
                    return 'Tumor Markers'
                if any(x in test_name_lower for x in ['culture', 'hvs', 'microbiology', 'fungal', 'koh', 'sputum', 'afb', 'semen']):
                    return 'Microbiology'
                if any(x in test_name_lower for x in ['urine routine', 'stool routine', 'urine r/e', 'stool r/e', 'microscopy']):
                    return 'Urinalysis'
                if any(x in test_name_lower for x in ['malaria', 'parasite', 'skin snip']):
                    return 'Parasitology'
                return 'Immunology'

            for row in range(9, ws.max_row + 1):
                col_a = ws.cell(row, 1).value
                col_b = ws.cell(row, 2).value
                col_d = ws.cell(row, 4).value
                col_e = ws.cell(row, 5).value

                if not col_a and not col_b:
                    continue

                # Strip whitespace
                col_a_str = str(col_a).strip() if col_a else ''
                col_b_str = str(col_b).strip() if col_b else ''

                # Check if it's a test/category header row (Col A has prefix e.g., "1. Liver Function Test")
                if col_a and not col_b:
                    # Clean up prefix like "1. ", "35. ", etc.
                    clean_name = re.sub(r'^\d+\.\s*', '', col_a_str).strip()
                    # Skip noise headers and notes
                    if "additional routine tests" in clean_name.lower() or "patient & sample" in clean_name.lower() or clean_name.lower().startswith("note:") or clean_name.lower().startswith("legend:") or len(clean_name) > 100:
                        continue
                    
                    clean_name = clean_name[:140]
                    cat_name = get_category_for_test(clean_name)
                    category, _ = TestCategory.objects.get_or_create(
                        name=cat_name,
                        defaults={'description': f"Standard {cat_name} tests"}
                    )
                    
                    # Create Test under this category
                    code = clean_name.split('(')[-1].replace(')', '').strip().upper() if '(' in clean_name else clean_name[:5].upper().replace(' ', '')
                    # clean up code formatting
                    code = re.sub(r'[^A-Z0-9]', '', code)
                    code = code[:15]
                    if not code:
                        code = f"TST{row}"
                    
                    # Create/get Test
                    current_test, created = Test.objects.get_or_create(
                        code=code,
                        defaults={
                            'category': category,
                            'name': clean_name,
                            'price': 40.00, # default price
                            'turnaround_time_hours': 4,
                            'normal_range': 'See parameters',
                            'units': 'n/a',
                            'sample_type': 'Whole Blood' if cat_name == 'Haematology' else 'Serum'
                        }
                    )
                    self.stdout.write(f"Seeding Test: {current_test.name} ({current_test.code})")
                
                # Check if it's a parameter row
                elif col_a and col_b and current_test:
                    # Parameter name
                    param_name = col_b_str
                    normal_range = str(col_e).strip() if col_e else 'Normal'
                    units = str(col_d).strip() if col_d else 'n/a'
                    
                    # Check and create parameter
                    TestParameter.objects.get_or_create(
                        test=current_test,
                        name=param_name,
                        defaults={
                            'normal_range': normal_range,
                            'units': units,
                            'display_order': row
                        }
                    )
        else:
            self.stdout.write(self.style.WARNING("Lab_Results_Template.xlsx not found, skipping Excel-based seeding."))

        # Seed Phase 2 Supplier and Reagents
        self.stdout.write('Seeding Phase 2 inventory...')
        from inventory.models import Supplier, Reagent
        supplier, _ = Supplier.objects.get_or_create(
            name='Global Reagents Ltd',
            defaults={
                'contact_name': 'John Doe',
                'phone': '+1 555-0192',
                'email': 'sales@globalreagents.com',
                'address': '123 Bio Drive, Suite 100'
            }
        )

        reagents_data = [
            {'name': 'CBC Lysing Reagent', 'code': 'RG-CBC-LYSE', 'unit': 'Vial', 'min_stock_level': 5.0, 'current_quantity': 15.0},
            {'name': 'LFT Calibration Serum', 'code': 'RG-LFT-CAL', 'unit': 'Kit', 'min_stock_level': 2.0, 'current_quantity': 8.0},
            {'name': 'Urine Strip 10 Parameters', 'code': 'RG-UA-STRIP', 'unit': 'Box', 'min_stock_level': 10.0, 'current_quantity': 3.0}, # Low stock
        ]
        for rd in reagents_data:
            Reagent.objects.get_or_create(
                code=rd['code'],
                defaults={
                    'name': rd['name'],
                    'unit': rd['unit'],
                    'min_stock_level': rd['min_stock_level'],
                    'current_quantity': rd['current_quantity'],
                    'supplier': supplier
                }
            )

        # Seed Phase 2 Equipment
        self.stdout.write('Seeding Phase 2 equipment...')
        from equipment.models import Equipment
        Equipment.objects.get_or_create(
            serial_number='SN-XM200-9304',
            defaults={
                'name': 'Sysmex CBC Analyzer',
                'model': 'XN-1000',
                'manufacturer': 'Sysmex Corporation',
                'status': 'active'
            }
        )
        Equipment.objects.get_or_create(
            serial_number='SN-COB400-8812',
            defaults={
                'name': 'Cobas Biochemistry Analyzer',
                'model': 'c311',
                'manufacturer': 'Roche Diagnostics',
                'status': 'active'
            }
        )

        # Seed Phase 2 Quality Control
        self.stdout.write('Seeding Quality Control parameters from Excel...')
        from quality.models import QCControl, QCRun
        import os
        import datetime
        from django.conf import settings
        from decimal import Decimal
        
        admin_user = User.objects.filter(role=User.ROLE_ADMINISTRATOR).first()
        qc_excel_path = os.path.join(settings.BASE_DIR, "Laboratory_Inventory_QC_Equipment.xlsx")
        if not os.path.exists(qc_excel_path):
            qc_excel_path = r"c:\Users\HP\Desktop\LMS\Laboratory_Inventory_QC_Equipment.xlsx"

        if os.path.exists(qc_excel_path):
            wb = openpyxl.load_workbook(qc_excel_path)
            if 'Quality Control' in wb.sheetnames:
                ws = wb['Quality Control']
                current_category = 'General'
                for row in range(4, ws.max_row + 1):
                    col_b = ws.cell(row, 2).value
                    if not col_b:
                        col_a = ws.cell(row, 1).value
                        if col_a:
                            clean_cat = re.sub(r'^\d+\.\s*', '', str(col_a)).strip()
                            if not clean_cat.lower().startswith("legend:") and len(clean_cat) <= 100:
                                current_category = clean_cat
                        continue
                    col_b_str = str(col_b).strip()
                    if col_b_str.lower().startswith("legend:"):
                        continue
                    
                    col_a = ws.cell(row, 1).value
                    col_c = ws.cell(row, 3).value
                    col_d = ws.cell(row, 4).value
                    col_e = ws.cell(row, 5).value
                    col_f = ws.cell(row, 6).value
                    col_g = ws.cell(row, 7).value

                    control_name = col_b_str
                    level = str(col_c).strip() if col_c else ''
                    
                    if level and level.upper() != 'N/A':
                        name = f"{control_name} - {level}"
                    else:
                        name = control_name

                    # Lot Number
                    if col_d:
                        lot_number = str(col_d).strip()
                    else:
                        control_words = [w for w in re.split(r'[^a-zA-Z0-9]', control_name) if w]
                        control_code = "".join(w[0].upper() for w in control_words)[:5]
                        level_words = [w for w in re.split(r'[^a-zA-Z0-9]', level) if w]
                        level_code = "".join(w[0].upper() for w in level_words)[:5] if level_words else 'NA'
                        lot_number = f"LOT-{control_code}-{level_code}-{row}"

                    # Target Value
                    if col_e is not None:
                        target_value = str(col_e).strip()
                    else:
                        level_lower = level.lower()
                        if 'negative' in level_lower:
                            target_value = "Negative"
                        elif 'positive' in level_lower:
                            target_value = "Positive"
                        elif 'level 1' in level_lower or 'low' in level_lower or 'normal' in level_lower:
                            target_value = "5.0"
                        elif 'level 2' in level_lower or 'elevated' in level_lower or 'abnormal' in level_lower:
                            target_value = "15.0"
                        elif 'level 3' in level_lower or 'high' in level_lower:
                            target_value = "25.0"
                        else:
                            target_value = "Normal"

                    # Standard Deviation
                    if col_f is not None:
                        sd_str = str(col_f).strip().replace('+/-', '').replace('%', '')
                        try:
                            standard_deviation = Decimal(sd_str)
                        except Exception:
                            standard_deviation = Decimal('0.0000')
                    else:
                        try:
                            target_val_float = float(target_value)
                            standard_deviation = Decimal(f"{target_val_float * 0.1:.4f}")
                        except ValueError:
                            standard_deviation = Decimal('0.0000')

                    # Expiry Date
                    if col_g is not None:
                        if isinstance(col_g, (datetime.date, datetime.datetime)):
                            expiry_date = col_g.date() if isinstance(col_g, datetime.datetime) else col_g
                        else:
                            try:
                                expiry_date = datetime.datetime.strptime(str(col_g).strip(), "%Y-%m-%d").date()
                            except ValueError:
                                expiry_date = datetime.date(2027, 12, 31)
                    else:
                        expiry_date = datetime.date(2027, 7, 21)

                    qc_control, _ = QCControl.objects.update_or_create(
                        lot_number=lot_number,
                        defaults={
                            'name': name[:100],
                            'category': current_category[:100],
                            'expiry_date': expiry_date,
                            'target_value': target_value,
                            'standard_deviation': standard_deviation,
                            'created_by': admin_user,
                            'updated_by': admin_user,
                        }
                    )
                    
                    if not QCRun.objects.filter(control=qc_control).exists():
                        QCRun.objects.create(
                            control=qc_control,
                            value=target_value,
                            status='pass',
                            notes='Initial calibration pass on load.',
                            created_by=admin_user,
                            updated_by=admin_user,
                        )
            else:
                self.stdout.write(self.style.WARNING("Quality Control sheet not found in Excel."))
        else:
            self.stdout.write(self.style.WARNING("Laboratory_Inventory_QC_Equipment.xlsx not found, skipping Excel-based QC seeding."))

        # Seed Phase 2 Referrals Partner
        self.stdout.write('Seeding Phase 2 referral partners...')
        from referrals.models import ReferralPartner
        ReferralPartner.objects.get_or_create(
            name='St. Jude Reference Laboratory',
            defaults={
                'contact_number': '+1 555-9302',
                'email': 'referrals@stjude.org',
                'address': '456 Medical Parkway, Memphis'
            }
        )

        # Seed Phase 3 Demo Patient User Account
        self.stdout.write('Seeding Phase 3 patient user account...')
        from patients.models import Patient
        demo_patient = Patient.objects.filter(first_name='Alice', last_name='Wonderland').first()
        if not demo_patient:
            demo_patient = Patient.objects.create(
                first_name='Alice',
                last_name='Wonderland',
                gender='F',
                date_of_birth='1995-04-12',
                phone_number='+1555-0143'
            )
        if not User.objects.filter(username='patient').exists():
            User.objects.create_user(
                username='patient',
                email='patient@lms.com',
                password='Password123!',
                role=User.ROLE_PATIENT,
                first_name='Alice',
                last_name='Wonderland',
                patient=demo_patient
            )
            self.stdout.write(self.style.SUCCESS("Created patient user: patient"))

        self.stdout.write(self.style.SUCCESS('Database seeding completed successfully!'))
