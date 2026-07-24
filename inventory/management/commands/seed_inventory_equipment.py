import os
import re
import datetime
from decimal import Decimal
import openpyxl

from django.core.management.base import BaseCommand
from django.conf import settings
from equipment.models import Equipment
from inventory.models import Supplier, Reagent


class Command(BaseCommand):
    help = 'Seeds analyzers, equipment, inventory, and reagents from the Excel master list'

    def handle(self, *args, **options):
        self.stdout.write(self.style.NOTICE('Starting Excel-based database seeding...'))

        # Locate the Excel file
        excel_path = os.path.join(settings.BASE_DIR, "Laboratory_Inventory_QC_Equipment.xlsx")
        if not os.path.exists(excel_path):
            excel_path = r"c:\Users\HP\Desktop\LMS\Laboratory_Inventory_QC_Equipment.xlsx"

        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR(f"Excel file not found at: {excel_path}"))
            return

        try:
            wb = openpyxl.load_workbook(excel_path)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Failed to load workbook: {str(e)}"))
            return

        # Initialize defaults
        today = datetime.date.today()
        default_purchase_date = today - datetime.timedelta(days=365)
        default_last_calibration = today - datetime.timedelta(days=90)
        default_next_calibration = today + datetime.timedelta(days=92)
        default_expiry_date = datetime.date(2027, 12, 31)

        # Ensure we have a default supplier for Reagents
        supplier, _ = Supplier.objects.get_or_create(
            name='Global Reagents Ltd',
            defaults={
                'contact_name': 'John Doe',
                'phone': '+1 555-0192',
                'email': 'sales@globalreagents.com',
                'address': '123 Bio Drive, Suite 100'
            }
        )

        # ----------------------------------------------------
        # 1. Seed Analyzers & Equipment
        # ----------------------------------------------------
        if 'Analysers & Equipment' in wb.sheetnames:
            self.stdout.write(self.style.NOTICE('Seeding Analyzers & Equipment sheet...'))
            ws = wb['Analysers & Equipment']
            equipment_seeded = 0

            # Loop starting from row 4 (skipping headers/titles)
            for row in range(4, ws.max_row + 1):
                col_c = ws.cell(row, 3).value  # Serial Number
                if not col_c:
                    continue

                serial_number = str(col_c).strip()
                if serial_number.lower().startswith("legend:") or not serial_number:
                    continue

                col_a = ws.cell(row, 1).value  # Name
                col_b = ws.cell(row, 2).value  # Model
                col_d = ws.cell(row, 4).value  # Manufacturer

                name = str(col_a).strip() if col_a else "Unknown Equipment"
                model = str(col_b).strip() if col_b else "N/A"
                manufacturer = str(col_d).strip() if col_d else "Unknown"

                # Check if it looks like a section header row
                if not col_b and not col_d:
                    # E.g. 'MICROBIOLOGY & MOLECULAR' or 'GENERAL LABORATORY EQUIPMENT'
                    continue

                eq, created = Equipment.objects.update_or_create(
                    serial_number=serial_number,
                    defaults={
                        'name': name[:150],
                        'model': model[:100],
                        'manufacturer': manufacturer[:100],
                        'purchase_date': default_purchase_date,
                        'last_calibration_date': default_last_calibration,
                        'next_calibration_date': default_next_calibration,
                        'status': Equipment.STATUS_ACTIVE
                    }
                )
                if created:
                    equipment_seeded += 1
                    self.stdout.write(f"  Created Equipment: {eq.name} ({eq.serial_number})")
                else:
                    self.stdout.write(f"  Updated Equipment: {eq.name} ({eq.serial_number})")

            self.stdout.write(self.style.SUCCESS(f"Finished seeding Analyzers & Equipment ({equipment_seeded} new equipment created)."))
        else:
            self.stdout.write(self.style.WARNING("Sheet 'Analysers & Equipment' not found in workbook."))

        # ----------------------------------------------------
        # 2. Seed Reagents & Consumables
        # ----------------------------------------------------
        if 'Inventory' in wb.sheetnames:
            self.stdout.write(self.style.NOTICE('Seeding Reagents & Inventory sheet...'))
            ws = wb['Inventory']
            reagents_seeded = 0

            # Loop starting from row 4
            for row in range(4, ws.max_row + 1):
                col_c = ws.cell(row, 3).value  # Code
                if not col_c:
                    continue

                code = str(col_c).strip()
                if code.lower().startswith("legend:") or not code:
                    continue

                col_a = ws.cell(row, 1).value  # Category
                col_b = ws.cell(row, 2).value  # Product Name
                col_d = ws.cell(row, 4).value  # Unit
                col_e = ws.cell(row, 5).value  # Minimum Level
                col_f = ws.cell(row, 6).value  # Expiry Date

                category = str(col_a).strip() if col_a else "Reagent"
                name = str(col_b).strip() if col_b else "Unknown Item"
                unit = str(col_d).strip() if col_d else "Kit"

                # Parse minimum level
                try:
                    min_stock_level = Decimal(str(col_e).strip()) if col_e is not None else Decimal('0.00')
                except Exception:
                    min_stock_level = Decimal('0.00')

                # Calculate default current quantity (min_stock_level * 2 or 10.0 if 0)
                current_quantity = min_stock_level * Decimal('2.0') if min_stock_level > 0 else Decimal('10.0')

                # Parse expiry date
                expiry_date = default_expiry_date
                if col_f is not None:
                    if isinstance(col_f, (datetime.date, datetime.datetime)):
                        expiry_date = col_f.date() if isinstance(col_f, datetime.datetime) else col_f
                    else:
                        # Attempt string parse
                        try:
                            expiry_date = datetime.datetime.strptime(str(col_f).strip(), "%Y-%m-%d").date()
                        except ValueError:
                            pass

                reagent, created = Reagent.objects.update_or_create(
                    code=code,
                    defaults={
                        'name': name[:150],
                        'unit': unit[:30],
                        'min_stock_level': min_stock_level,
                        'current_quantity': current_quantity,
                        'expiry_date': expiry_date,
                        'supplier': supplier,
                        'description': f"Imported {category} from Excel list"
                    }
                )
                if created:
                    reagents_seeded += 1
                    self.stdout.write(f"  Created Reagent: {reagent.name} ({reagent.code})")
                else:
                    self.stdout.write(f"  Updated Reagent: {reagent.name} ({reagent.code})")

            self.stdout.write(self.style.SUCCESS(f"Finished seeding Inventory & Reagents ({reagents_seeded} new reagents created)."))
        else:
            self.stdout.write(self.style.WARNING("Sheet 'Inventory' not found in workbook."))

        self.stdout.write(self.style.SUCCESS('Excel database seeding completed!'))
